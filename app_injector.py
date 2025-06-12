import streamlit as st
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill, Font, Alignment
from copy import copy
import tempfile
import os

def inject_monthly_sheets_with_formatting(template_path, user_file_path, output_path):
    template_wb = load_workbook(template_path, data_only=False)
    user_wb = load_workbook(user_file_path)

    monthly_sheets = [name for name in template_wb.sheetnames if name.startswith("2025")]

    for sheet in monthly_sheets:
        if sheet in user_wb.sheetnames:
            del user_wb[sheet]

    for sheet_name in monthly_sheets:
        source_ws = template_wb[sheet_name]
        new_ws = user_wb.create_sheet(title=sheet_name)

        for row in source_ws.iter_rows():
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        for row_idx, row_dim in source_ws.row_dimensions.items():
            new_ws.row_dimensions[row_idx].height = row_dim.height
        for col_letter, col_dim in source_ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = col_dim.width
        for merged_range in source_ws.merged_cells.ranges:
            new_ws.merge_cells(str(merged_range))

    user_wb.save(output_path)

# Streamlit UI
st.title("📊 Monthly Sheet Injector with Formatting")

st.markdown("""
Upload a **template Excel file** (with monthly sheets starting with "2025") and a **user Excel file**.
This app will inject the monthly sheets into the user file, preserving full formatting.
""")

template_file = st.file_uploader("Upload Template Excel File", type=["xlsx"])
user_file = st.file_uploader("Upload User Excel File", type=["xlsx"])

if template_file and user_file:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_template,          tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_user,          tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_output:

        tmp_template.write(template_file.read())
        tmp_user.write(user_file.read())
        tmp_template.flush()
        tmp_user.flush()

        inject_monthly_sheets_with_formatting(tmp_template.name, tmp_user.name, tmp_output.name)

        with open(tmp_output.name, "rb") as result:
            st.success("✅ Processing complete! Download your updated file below.")
            st.download_button("📥 Download Updated Excel", result, file_name="updated_user_file.xlsx")
